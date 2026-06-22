from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os
print("=====Добро пожаловать в Exel_py_Worker")
table_name = str(input("Введите имя вашей таблице Exel: "))
wb = Workbook()
sheet = wb.active
running = True
while running:
    print("Имя сохранено")
    activity = int(input(
"""\nВыбирите действие которое вам нужно:
Добавить текст в таблицу - 1
Закрасить определенную ячейку - 2
Вывести таблицу - 3
Редактировать существующую таблицу - 4
Выйти - 0
: \n"""))
    if activity == 1:
        select = str(input("Введите ячейку в которую хотите записать текст(например: A1): "))
        print("Успешно")
        text_select = str(input(f"Введите текс который хотите записать в ячейке {select}: "))
        sheet[select] = text_select
    elif activity == 2:
        select_1 = str(input("Введите ячейку которую хотели бы закрасить(например: A1): "))
        select_color = str(input("Введите цвет который хотите покрасить ячейку(например: FFFFFF): "))
        color = PatternFill(start_color=select_color, end_color=select_color, fill_type="solid")
        sheet[select_1].fill = color
    elif activity == 3:
        selct_name = str(input("Введите имя файлы который хотите прочесть: "))
        wb_read = load_workbook(selct_name + ".xlsx")
        sheet_read = wb_read.active
        
        print("\n--- Содержимое ---")
        for row in sheet_read.iter_rows(values_only=True):
            print(row)
        print("------------------\n")
    elif activity == 4:
        name = str(input("Введите имя СУЩЕСТВУЕЩЕГО ФАЙЛА: "))
        file_name = name + ".xlsx"
        if os.path.exists(file_name):
            print("Файл найден")
            wb = load_workbook(file_name)
            sheet = wb.active
            table_name = name
        else:
            print(f"Файл {name} не найден.")
            wb = Workbook
    elif activity == 0:
        wb.save(table_name + ".xlsx")
        running = False