from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os

file_name = "Codes.xlsx"
wb = load_workbook(file_name)
sheet = wb.active
find = input(f"Введите строку которую хотите найти в {file_name}: ")
if find == sheet["A1"]:
    sheet["B1"] == find
    sheet["B1"].fill = "Yellow"
else:
    print("Строка не найдена!")